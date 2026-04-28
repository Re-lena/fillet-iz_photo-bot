import logging
from flask import Flask, request, jsonify
import cv2
import numpy as np
from PIL import Image
import io
import os
import requests
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
from time import sleep

logging.basicConfig(level=logging.INFO)

TOKEN = os.environ.get('TELEGRAM_BOT_TOKEN')
if not TOKEN:
    raise ValueError("TELEGRAM_BOT_TOKEN not set")

app = Flask(__name__)

# ---------- Константы ----------
DEFAULT_CELLS = 50
MIN_CELLS = 5
MAX_CELLS = 200
MAX_PHOTO_BYTES = 10 * 1024 * 1024   # 10 МБ

ADAPTIVE_BLOCK_SIZE = 11
ADAPTIVE_C = 2
MORPH_KERNEL_SIZE = 3                # увеличено с 2 до 3 для лучшего закрытия дырок
CLAHE_CLIP_LIMIT = 2.0
CLAHE_TILE_SIZE = 8

user_settings = {}

# ---------- Повторные попытки ----------
def send_with_retry(send_func, retries=3, delay=1):
    for attempt in range(retries):
        try:
            return send_func()
        except Exception as e:
            logging.warning(f"Attempt {attempt+1} failed: {e}")
            if attempt < retries - 1:
                sleep(delay)
            else:
                raise

# ---------- Отправка файлов и сообщений ----------
def send_document(chat_id, file_bytes, filename, caption=""):
    url = f"https://api.telegram.org/bot{TOKEN}/sendDocument"
    files = {'document': (filename, file_bytes, 'application/octet-stream')}
    data = {'chat_id': chat_id, 'caption': caption}
    return send_with_retry(lambda: requests.post(url, files=files, data=data))

def send_photo(chat_id, photo_bytes, caption=""):
    url = f"https://api.telegram.org/bot{TOKEN}/sendPhoto"
    files = {'photo': ('scheme.png', photo_bytes, 'image/png')}
    data = {'chat_id': chat_id, 'caption': caption}
    return send_with_retry(lambda: requests.post(url, files=files, data=data))

def send_message(chat_id, text):
    url = f"https://api.telegram.org/bot{TOKEN}/sendMessage"
    payload = {'chat_id': chat_id, 'text': text}
    return requests.post(url, json=payload)

def send_menu_keyboard(chat_id, text):
    url = f"https://api.telegram.org/bot{TOKEN}/sendMessage"
    keyboard = {
        "keyboard": [
            ["🐭 Маленький", "🐰 Средний"],
            ["🐘 Большой", "📏 Свой размер"],
            ["❓ Помощь"]
        ],
        "resize_keyboard": True,
        "one_time_keyboard": False
    }
    payload = {
        'chat_id': chat_id,
        'text': text,
        'reply_markup': keyboard
    }
    return requests.post(url, json=payload)

# ---------- Обработка изображения (оптимизированная) ----------
def process_image_to_matrix(image_bytes, target_cells):
    image = Image.open(io.BytesIO(image_bytes))
    img_cv = cv2.cvtColor(np.array(image), cv2.COLOR_RGB2BGR)
    gray = cv2.cvtColor(img_cv, cv2.COLOR_BGR2GRAY)

    clahe = cv2.createCLAHE(clipLimit=CLAHE_CLIP_LIMIT, tileGridSize=(CLAHE_TILE_SIZE, CLAHE_TILE_SIZE))
    gray = clahe.apply(gray)

    binary = cv2.adaptiveThreshold(gray, 255,
                                   cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                                   cv2.THRESH_BINARY_INV,
                                   ADAPTIVE_BLOCK_SIZE, ADAPTIVE_C)

    kernel = np.ones((MORPH_KERNEL_SIZE, MORPH_KERNEL_SIZE), np.uint8)
    cleaned = cv2.morphologyEx(binary, cv2.MORPH_CLOSE, kernel)

    h, w = cleaned.shape
    target_cols = max(MIN_CELLS, min(target_cells, MAX_CELLS))
    target_rows = max(1, int(target_cols * (h / w)))
    if target_rows > 200:
        target_rows = 200
        target_cols = int(target_rows * (w / h))

    small = cv2.resize(cleaned, (target_cols, target_rows), interpolation=cv2.INTER_AREA)

    matrix = []
    for row in range(target_rows):
        row_str = ''.join('1' if small[row, col] == 255 else '0' for col in range(target_cols))
        matrix.append(row_str)

    scheme_pil = Image.fromarray(small, mode='L')
    return scheme_pil, matrix

# ---------- Excel с полным форматированием (исправлено) ----------
def generate_excel_bytes(matrix):
    wb = Workbook()
    ws = wb.active
    ws.title = "Scheme"

    black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    white_font = Font(color="FFFFFF")
    black_font = Font(color="000000")
    center_align = Alignment(horizontal="center", vertical="center")

    rows = len(matrix)
    cols = len(matrix[0]) if rows > 0 else 0

    # Заголовки столбцов
    for col_idx in range(1, cols + 1):
        cell = ws.cell(row=1, column=col_idx + 1)
        cell.value = col_idx
        cell.alignment = center_align

    # Заголовки строк
    for row_idx in range(1, rows + 1):
        cell = ws.cell(row=row_idx + 1, column=1)
        cell.value = row_idx
        cell.alignment = center_align

    # Заливка схемы
    for i, row_str in enumerate(matrix):
        for j, ch in enumerate(row_str):
            cell = ws.cell(row=i + 2, column=j + 2)
            cell.value = int(ch)
            cell.alignment = center_align
            if ch == '1':
                cell.fill = black_fill
                cell.font = white_font
            else:
                cell.fill = white_fill
                cell.font = black_font

    # Квадратные ячейки
    for col in range(2, cols + 2):
        ws.column_dimensions[get_column_letter(col)].width = 3
    for row in range(2, rows + 2):
        ws.row_dimensions[row].height = 15

    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer

# ---------- Текстовое описание (без изменений) ----------
def generate_description_txt(matrix):
    lines = []
    for idx, row_str in enumerate(matrix, start=1):
        if not row_str:
            lines.append(f"Ряд {idx}: пустая строка")
            continue
        groups = []
        current = row_str[0]
        count = 1
        for ch in row_str[1:]:
            if ch == current:
                count += 1
            else:
                groups.append(f"{count} {'пустых' if current == '0' else 'заполненных'}")
                current = ch
                count = 1
        groups.append(f"{count} {'пустых' if current == '0' else 'заполненных'}")
        lines.append(f"Ряд {idx}: " + ", ".join(groups))
    return "\n".join(lines)

# ---------- Вебхук ----------
@app.route('/webhook', methods=['POST'])
def webhook():
    try:
        update = request.get_json()
        if not update or 'message' not in update:
            return jsonify({'status': 'ok'})

        message = update['message']
        chat_id = message['chat']['id']
        text = message.get('text', '')

        # Преобразование кнопок
        if text == "🐭 Маленький":
            text = "/small"
        elif text == "🐰 Средний":
            text = "/medium"
        elif text == "🐘 Большой":
            text = "/big"
        elif text == "📏 Свой размер":
            text = "/cells"
        elif text == "❓ Помощь":
            text = "/help"

        # Команды
        if text.startswith('/big'):
            user_settings[chat_id] = 50
            send_message(chat_id, "✅ Установлен большой размер схемы (~50 ячеек по ширине).")
            return jsonify({'status': 'ok'})

        if text.startswith('/medium'):
            user_settings[chat_id] = 25
            send_message(chat_id, "✅ Установлен средний размер схемы (~25 ячеек по ширине).")
            return jsonify({'status': 'ok'})

        if text.startswith('/small'):
            user_settings[chat_id] = 15
            send_message(chat_id, "✅ Установлен маленький размер схемы (~15 ячеек по ширине).")
            return jsonify({'status': 'ok'})

        if text.startswith('/cells'):
            parts = text.split()
            if len(parts) == 2:
                try:
                    val = int(parts[1])
                    if MIN_CELLS <= val <= MAX_CELLS:
                        user_settings[chat_id] = val
                        send_message(chat_id, f"✅ Количество ячеек по ширине установлено: {val}.")
                    else:
                        send_message(chat_id, f"❌ Введите число от {MIN_CELLS} до {MAX_CELLS}.")
                except:
                    send_message(chat_id, "❌ Используйте: /cells <число>")
            else:
                send_message(chat_id, f"Введите количество ячеек по ширине числом от {MIN_CELLS} до {MAX_CELLS}. Например: /cells 30")
            return jsonify({'status': 'ok'})

        if text.startswith('/start') or text.startswith('/help'):
            help_text = "🧶 Выберите размер будущего изделия:"
            send_menu_keyboard(chat_id, help_text)
            return jsonify({'status': 'ok'})

        # Если нет фото
        if 'photo' not in message:
            return jsonify({'status': 'ok'})

        # Проверка размера
        photo_obj = message['photo'][-1]
        if photo_obj.get('file_size', 0) > MAX_PHOTO_BYTES:
            send_message(chat_id, "❌ Фото слишком большое (больше 10 МБ). Пожалуйста, отправьте изображение меньшего размера.")
            return jsonify({'status': 'ok'})

        target_cells = user_settings.get(chat_id, DEFAULT_CELLS)

        # Скачиваем фото
        file_id = photo_obj['file_id']
        file_info = requests.get(f"https://api.telegram.org/bot{TOKEN}/getFile?file_id={file_id}").json()
        if not file_info['ok']:
            raise Exception("Не удалось получить информацию о файле")

        file_path = file_info['result']['file_path']
        photo_bytes = requests.get(f"https://api.telegram.org/file/bot{TOKEN}/{file_path}").content

        # Генерация
        scheme_image, matrix = process_image_to_matrix(photo_bytes, target_cells)

        # PNG
        png_buffer = io.BytesIO()
        scheme_image.save(png_buffer, format='PNG', optimize=True)
        png_buffer.seek(0)
        send_photo(chat_id, png_buffer, f"📐 Схема (ширина {len(matrix[0])} ячеек, высота {len(matrix)})")

        # Excel (теперь с заливкой)
        excel_buffer = generate_excel_bytes(matrix)
        send_document(chat_id, excel_buffer, "scheme.xlsx", "📊 Excel-схема: 0=белый, 1=чёрный")

        # Текст
        description = generate_description_txt(matrix)
        txt_buffer = io.BytesIO(description.encode('utf-8'))
        txt_buffer.seek(0)
        send_document(chat_id, txt_buffer, "description.txt", "📄 Текстовое описание рядов")

        return jsonify({'status': 'ok'})
    except Exception as e:
        logging.error(f"Ошибка в вебхуке: {e}", exc_info=True)
        # Безопасно отправляем сообщение об ошибке (если chat_id существует)
        try:
            if 'chat_id' in locals():
                send_message(chat_id, "❌ Произошла ошибка при обработке. Попробуйте другое фото или размер.")
        except:
            pass
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/')
def index():
    return "Бот для филейного вязания работает. Используйте Telegram."

if __name__ == '__main__':
    app.run()
